from flask import Blueprint, render_template, request, redirect, url_for, flash, Response
from app.database import get_account_by_name, get_all_accounts
from app.services.analytics import compute_risk_all

calendario_bp = Blueprint("calendario", __name__)


def _group_by_urgency(items: list) -> dict:
    """
    Groups items into urgency buckets based on days_until_rupture.
    Returns ordered dict: HOJE / EM 5 DIAS / EM 10 DIAS / EM 20 DIAS / EM 30 DIAS
    """
    groups = {
        "HOJE":       [],
        "EM 5 DIAS":  [],
        "EM 10 DIAS": [],
        "EM 20 DIAS": [],
        "EM 30 DIAS": [],
    }
    for item in items:
        d = item["risk"].get("days_until_rupture")
        level = item["risk"].get("level", "ESTÁVEL")

        if level == "RUPTURA" or (d is not None and d <= 0):
            groups["HOJE"].append(item)
        elif d is not None and d <= 5:
            groups["EM 5 DIAS"].append(item)
        elif d is not None and d <= 10:
            groups["EM 10 DIAS"].append(item)
        elif d is not None and d <= 20:
            groups["EM 20 DIAS"].append(item)
        else:
            groups["EM 30 DIAS"].append(item)

    return groups


@calendario_bp.route("/calendario/<seller_id>")
def calendario(seller_id: str):
    accounts   = get_all_accounts()
    authorized = [a for a in accounts if a.get("access_token")]

    account = next((a for a in authorized if str(a["seller_id"]) == str(seller_id)), None)
    if not account:
        flash("Conta não encontrada ou não autorizada.", "error")
        return redirect(url_for("web.index"))

    risk_items = compute_risk_all(seller_id)
    groups     = _group_by_urgency(risk_items)

    total_at_risk = len(risk_items)
    ruptura_count = len(groups["HOJE"])

    return render_template(
        "calendario.html",
        account        = account,
        authorized     = authorized,
        groups         = groups,
        total_at_risk  = total_at_risk,
        ruptura_count  = ruptura_count,
        seller_id      = seller_id,
    )


@calendario_bp.route("/calendario/<seller_id>/export")
def calendario_export(seller_id: str):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        flash("openpyxl não instalado. Execute: pip install openpyxl", "error")
        return redirect(url_for("calendario.calendario", seller_id=seller_id))

    accounts   = get_all_accounts()
    authorized = [a for a in accounts if a.get("access_token")]
    account    = next((a for a in authorized if str(a["seller_id"]) == str(seller_id)), None)
    if not account:
        flash("Conta não encontrada.", "error")
        return redirect(url_for("web.index"))

    risk_items = compute_risk_all(seller_id)
    groups     = _group_by_urgency(risk_items)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Calendário de Reposição"

    # Header
    headers = ["Urgência", "Item ID", "Título", "Estoque", "Giro 7d", "Giro 30d",
               "Dias até Ruptura", "Risco Score", "Nível", "Envio Sugerido", "Prazo Envio"]
    header_fill = PatternFill("solid", fgColor="1a1a1a")
    header_font = Font(bold=True, color="F5A623")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    row_num = 2
    group_colors = {
        "HOJE":       "F44336",
        "EM 5 DIAS":  "FF9800",
        "EM 10 DIAS": "FFC107",
        "EM 20 DIAS": "8BC34A",
        "EM 30 DIAS": "4CAF50",
    }

    for group_name, items in groups.items():
        if not items:
            continue

        fill_color = group_colors.get(group_name, "FFFFFF")

        for item in items:
            risk    = item["risk"]
            induct  = item.get("induction", {})

            ws.cell(row=row_num, column=1, value=group_name).fill = PatternFill("solid", fgColor=fill_color)
            ws.cell(row=row_num, column=2, value=item.get("id", ""))
            ws.cell(row=row_num, column=3, value=item.get("title", ""))
            ws.cell(row=row_num, column=4, value=item.get("stock", 0) or 0)
            ws.cell(row=row_num, column=5, value=round(item.get("giro_7d", 0) or 0, 1))
            ws.cell(row=row_num, column=6, value=round(item.get("giro_30d", 0) or 0, 1))
            ws.cell(row=row_num, column=7, value=risk.get("days_until_rupture"))
            ws.cell(row=row_num, column=8, value=risk.get("score", 0))
            ws.cell(row=row_num, column=9, value=risk.get("level", ""))
            ws.cell(row=row_num, column=10, value=induct.get("qty", 0))
            ws.cell(row=row_num, column=11, value=induct.get("prazo_label", ""))

            row_num += 1

    # Column widths
    col_widths = [12, 16, 45, 10, 10, 10, 16, 12, 10, 14, 16]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    import io
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    import datetime as dt
    fname = f"calendario_{seller_id}_{dt.date.today().isoformat()}.xlsx"

    return Response(
        buf.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"}
    )
