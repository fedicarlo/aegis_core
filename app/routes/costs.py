import io
import json
from flask import (
    Blueprint, render_template, redirect, url_for,
    flash, request, send_file, jsonify,
)
from app.database import (
    get_account_by_seller_id, get_items_for_costs,
    save_cost, get_seller_defaults, save_seller_defaults,
    apply_defaults_to_all, init_costs_table, get_cost_history,
    init_nfe_tables, save_nfe_document, save_nfe_items,
    auto_reconcile_nfe_items, get_pending_nfe_items, reconcile_nfe_item,
    init_ml_fee_cache_table, get_cached_ml_fee, save_ml_fee_cache, update_ml_fee_rate,
)
from app.services.nfe_parser import parse_nfe_xml, NFeParseError
from app.services.meli_api import get_listing_fee
from app.services.meli_auth import get_valid_token

costs_bp = Blueprint("costs", __name__)


def _require_seller(seller_id: str):
    init_costs_table()
    init_nfe_tables()
    init_ml_fee_cache_table()
    return get_account_by_seller_id(seller_id)


# ── Interface principal ───────────────────────────────────────────────────────

@costs_bp.route("/custos/<seller_id>")
def custos(seller_id):
    account = _require_seller(seller_id)
    if not account:
        flash("Conta não encontrada.", "error")
        return redirect(url_for("web.index"))

    items    = get_items_for_costs(seller_id)
    defaults = get_seller_defaults(seller_id)

    sem_custo = sum(1 for i in items if i["unit_cost"] is None)
    marcas    = sorted({i["marca"] for i in items if i.get("marca")})

    return render_template(
        "custos.html",
        account   = account,
        items     = items,
        defaults  = defaults,
        seller_id = seller_id,
        sem_custo = sem_custo,
        marcas    = marcas,
    )


# ── Salvar linha individual (mantido para compatibilidade) ────────────────────

@costs_bp.route("/custos/<seller_id>/save", methods=["POST"])
def save_cost_item(seller_id):
    account = _require_seller(seller_id)
    if not account:
        flash("Conta não encontrada.", "error")
        return redirect(url_for("web.index"))

    item_id       = request.form.get("item_id", "").strip()
    unit_cost     = float(request.form.get("unit_cost") or 0)
    tax_rate_pct  = float(request.form.get("tax_rate") or 0)
    ml_fee_pct    = float(request.form.get("ml_fee_rate") or 14)
    shipping_cost = float(request.form.get("shipping_cost") or 0)
    marca         = request.form.get("marca", "").strip()

    if not item_id:
        flash("item_id inválido.", "error")
        return redirect(url_for("costs.custos", seller_id=seller_id))

    save_cost(
        seller_id, item_id, None,
        unit_cost, tax_rate_pct / 100.0, ml_fee_pct / 100.0,
        shipping_cost, marca,
    )
    flash("Custo salvo.", "success")
    return redirect(url_for("costs.custos", seller_id=seller_id))


# ── Salvar lote (JSON) ────────────────────────────────────────────────────────

@costs_bp.route("/custos/<seller_id>/save-batch", methods=["POST"])
def save_batch(seller_id):
    account = _require_seller(seller_id)
    if not account:
        return jsonify({"ok": False, "error": "Conta não encontrada"}), 404

    try:
        rows = request.get_json(force=True)
    except Exception:
        return jsonify({"ok": False, "error": "JSON inválido"}), 400

    if not isinstance(rows, list):
        return jsonify({"ok": False, "error": "Esperado array JSON"}), 400

    saved  = 0
    errors = []
    for i, row in enumerate(rows):
        try:
            item_id       = str(row["item_id"]).strip()
            unit_cost     = float(row.get("unit_cost") or 0)
            tax_rate_pct  = float(row.get("tax_rate") or 0)
            ml_fee_pct    = float(row.get("ml_fee_rate") or 14)
            shipping_cost = float(row.get("shipping_cost") or 0)
            marca         = str(row.get("marca") or "").strip()

            save_cost(
                seller_id, item_id, None,
                unit_cost, tax_rate_pct / 100.0, ml_fee_pct / 100.0,
                shipping_cost, marca,
            )
            saved += 1
        except Exception as e:
            errors.append({"row": i, "item_id": row.get("item_id"), "error": str(e)})

    return jsonify({"ok": True, "saved": saved, "errors": errors})


# ── Histórico de custo por item ───────────────────────────────────────────────

@costs_bp.route("/custos/<seller_id>/history/<item_id>")
def cost_history(seller_id, item_id):
    account = _require_seller(seller_id)
    if not account:
        return jsonify([]), 404

    import datetime as dt
    rows = get_cost_history(seller_id, item_id, limit=5)
    result = []
    for r in rows:
        ts = r["changed_at"]
        result.append({
            "old_cost":   r["old_cost"],
            "new_cost":   r["new_cost"],
            "changed_at": dt.datetime.fromtimestamp(ts).strftime("%d/%m/%Y %H:%M") if ts else "—",
        })
    return jsonify(result)


# ── Comissão ML sob demanda (com cache curto) ─────────────────────────────────

@costs_bp.route("/custos/<seller_id>/ml-fee/<item_id>")
def ml_fee(seller_id, item_id):
    account = _require_seller(seller_id)
    if not account:
        return jsonify({"ok": False, "error": "Conta não encontrada"}), 404

    cached = get_cached_ml_fee(seller_id, item_id, max_age_hours=36)
    if cached:
        return jsonify({
            "ok": True, "cached": True,
            "percentage_fee": cached["percentage_fee"],
            "fixed_fee": cached["fixed_fee"],
        })

    try:
        token = get_valid_token(account)
        fee   = get_listing_fee(token, item_id)
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 502

    save_ml_fee_cache(
        seller_id, item_id, fee["percentage_fee"], fee["fixed_fee"],
        fee["sale_fee_amount"], fee["price"],
    )
    update_ml_fee_rate(seller_id, item_id, "", fee["percentage_fee"] / 100.0)

    return jsonify({
        "ok": True, "cached": False,
        "percentage_fee": fee["percentage_fee"],
        "fixed_fee": fee["fixed_fee"],
    })


# ── Defaults por seller ───────────────────────────────────────────────────────

@costs_bp.route("/custos/<seller_id>/save-defaults", methods=["POST"])
def save_defaults(seller_id):
    account = _require_seller(seller_id)
    if not account:
        flash("Conta não encontrada.", "error")
        return redirect(url_for("web.index"))

    tax_rate_pct = float(request.form.get("tax_rate") or 0)
    ml_fee_pct   = float(request.form.get("ml_fee_rate") or 14)
    apply_all    = request.form.get("apply_all") == "1"

    save_seller_defaults(seller_id, tax_rate_pct / 100.0, ml_fee_pct / 100.0)

    if apply_all:
        apply_defaults_to_all(seller_id, tax_rate_pct / 100.0, ml_fee_pct / 100.0)
        flash(
            f"Defaults salvos e aplicados a todos os produtos "
            f"(alíquota {tax_rate_pct}%, comissão ML {ml_fee_pct}%).",
            "success",
        )
    else:
        flash(f"Defaults salvos (alíquota {tax_rate_pct}%, comissão ML {ml_fee_pct}%).", "success")

    return redirect(url_for("costs.custos", seller_id=seller_id))


# ── Template .xlsx para preenchimento ────────────────────────────────────────

@costs_bp.route("/custos/template/<seller_id>")
def template_costs(seller_id):
    account = _require_seller(seller_id)
    if not account:
        flash("Conta não encontrada.", "error")
        return redirect(url_for("web.index"))

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Protection, Border, Side
        from openpyxl.worksheet.datavalidation import DataValidation
        from openpyxl.utils import get_column_letter
    except ImportError:
        flash("openpyxl não instalado. Execute: pip install openpyxl", "error")
        return redirect(url_for("costs.custos", seller_id=seller_id))

    items = get_items_for_costs(seller_id)

    wb = Workbook()

    # ── Aba Custos ────────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Custos"

    headers = [
        "item_id", "titulo", "marca", "variacao",
        "preco_atual", "custo_nf", "aliquota_%", "comissao_ml_%", "frete_unit",
    ]
    # Col indices: A=item_id(lock), B=titulo(lock), C=marca(edit), D=variacao(lock),
    #              E=preco_atual(lock), F=custo_nf(edit), G=aliquota(edit), H=comissao(edit), I=frete(edit)

    lock_fill  = PatternFill("solid", fgColor="0d0d0d")
    lock_font  = Font(color="555555", size=10)
    edit_fill  = PatternFill("solid", fgColor="111a11")
    edit_font  = Font(color="e0e0e0", size=10)
    hdr_fill   = PatternFill("solid", fgColor="0d0d0d")
    hdr_font   = Font(bold=True, color="F5A623", size=10)
    inst_font  = Font(color="888888", size=10, italic=True)

    locked_cols = {1, 2, 4, 5}   # A, B, D, E — item_id, titulo, variacao, preco_atual
    editable_cols = {3, 6, 7, 8, 9}  # C, F, G, H, I

    # Header row
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font      = hdr_font
        cell.fill      = hdr_fill
        cell.alignment = Alignment(horizontal="center")
        cell.protection = Protection(locked=True)

    # Data rows
    for row_i, item in enumerate(items, 2):
        vals = [
            item["id"],
            item["title"],
            item.get("marca") or "",
            item.get("variation_id") or "",
            item.get("price") or 0,
            item.get("unit_cost") if item.get("unit_cost") is not None else "",
            round((item.get("tax_rate") or 0) * 100, 2),
            round((item.get("ml_fee_rate") or 0.14) * 100, 2),
            item.get("shipping_cost") or 0,
        ]
        for col_i, val in enumerate(vals, 1):
            cell = ws.cell(row=row_i, column=col_i, value=val)
            if col_i in locked_cols:
                cell.font       = lock_font
                cell.fill       = lock_fill
                cell.protection = Protection(locked=True)
            else:
                cell.font       = edit_font
                cell.fill       = edit_fill
                cell.protection = Protection(locked=False)
            cell.alignment = Alignment(horizontal="right" if col_i >= 4 else "left")

    # Data validation: numeric for custo/aliquota/comissao/frete (cols F-I)
    for col_letter in ("F", "G", "H", "I"):
        dv = DataValidation(
            type="decimal", operator="greaterThanOrEqual", formula1="0",
            allow_blank=True,
            error="Insira um número ≥ 0",
            errorTitle="Valor inválido",
            prompt="Insira um valor numérico ≥ 0",
            promptTitle="Valor esperado",
        )
        dv.sqref = f"{col_letter}2:{col_letter}10000"
        ws.add_data_validation(dv)

    # Column widths + freeze
    for col, w in zip("ABCDEFGHI", [18, 50, 18, 14, 12, 12, 12, 14, 10]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"

    # Sheet protection: allow editing unlocked cells only
    ws.protection.sheet     = True
    ws.protection.password  = ""          # sem senha — só impede edição acidental
    ws.protection.enable()

    # ── Aba Instruções ────────────────────────────────────────────────────────
    wi = wb.create_sheet("Instruções")
    wi.sheet_view.showGridLines = False
    instructions = [
        ("AEGIS — Planilha de Custos", Font(bold=True, color="F5A623", size=14)),
        ("", None),
        ("COLUNAS BLOQUEADAS (não editar):", Font(bold=True, color="888888", size=10)),
        ("  item_id    — ID do anúncio no Mercado Livre", None),
        ("  titulo     — Título do anúncio", None),
        ("  variacao   — ID de variação (se houver)", None),
        ("  preco_atual — Preço atual de venda", None),
        ("", None),
        ("COLUNAS EDITÁVEIS:", Font(bold=True, color="4caf50", size=10)),
        ("  marca         — Marca do produto (para filtros)", None),
        ("  custo_nf      — Custo da nota fiscal por unidade (R$)", None),
        ("  aliquota_%    — Alíquota de imposto sobre a venda (%)", None),
        ("  comissao_ml_% — Comissão do Mercado Livre (%)", None),
        ("  frete_unit    — Custo de frete unitário estimado (R$)", None),
        ("", None),
        ("COMO USAR:", Font(bold=True, color="888888", size=10)),
        ("  1. Preencha as colunas editáveis (fundo verde escuro)", None),
        ("  2. Deixe em branco qualquer campo sem informação", None),
        ("  3. Salve o arquivo como .xlsx e faça upload na página de Custos", None),
        ("  4. A importação valida todos os valores antes de salvar", None),
        ("  5. Erros serão listados linha a linha — corrija e reimporte", None),
    ]
    for r_i, (text, fnt) in enumerate(instructions, 1):
        cell = wi.cell(row=r_i, column=1, value=text)
        if fnt:
            cell.font = fnt
        else:
            cell.font = inst_font
    wi.column_dimensions["A"].width = 70

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    import datetime as dt
    fname = f"custos_template_{account['name']}_{dt.date.today().isoformat()}.xlsx"
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=fname,
    )


# ── Exportar planilha com dados atuais ────────────────────────────────────────

@costs_bp.route("/custos/export/<seller_id>")
def export_costs(seller_id):
    account = _require_seller(seller_id)
    if not account:
        flash("Conta não encontrada.", "error")
        return redirect(url_for("web.index"))

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        flash("openpyxl não instalado. Execute: pip install openpyxl", "error")
        return redirect(url_for("costs.custos", seller_id=seller_id))

    items = get_items_for_costs(seller_id)

    wb = Workbook()
    ws = wb.active
    ws.title = "Custos"

    headers = [
        "item_id", "titulo", "marca", "variacao",
        "preco_atual", "custo_nf", "aliquota_%",
        "comissao_ml_%", "frete_unit",
    ]
    ws.append(headers)

    hdr_fill = PatternFill("solid", fgColor="111111")
    for cell in ws[1]:
        cell.font      = Font(bold=True, color="F5A623", size=11)
        cell.fill      = hdr_fill
        cell.alignment = Alignment(horizontal="center")

    for item in items:
        ws.append([
            item["id"],
            item["title"],
            item.get("marca") or "",
            item.get("variation_id") or "",
            item.get("price") or "",
            item.get("unit_cost") if item.get("unit_cost") is not None else "",
            round((item.get("tax_rate") or 0) * 100, 2),
            round((item.get("ml_fee_rate") or 0.14) * 100, 2),
            item.get("shipping_cost") or 0,
        ])

    for col, width in zip("ABCDEFGHI", [18, 52, 18, 14, 14, 14, 14, 16, 12]):
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"custos_{account['name']}_{seller_id}.xlsx",
    )


# ── Importar planilha (validação completa antes de salvar) ────────────────────

@costs_bp.route("/custos/import/<seller_id>", methods=["POST"])
def import_costs(seller_id):
    account = _require_seller(seller_id)
    if not account:
        flash("Conta não encontrada.", "error")
        return redirect(url_for("costs.custos", seller_id=seller_id))

    try:
        from openpyxl import load_workbook
    except ImportError:
        flash("openpyxl não instalado. Execute: pip install openpyxl", "error")
        return redirect(url_for("costs.custos", seller_id=seller_id))

    file = request.files.get("file")
    if not file or not file.filename.endswith(".xlsx"):
        flash("Envie um arquivo .xlsx válido.", "error")
        return redirect(url_for("costs.custos", seller_id=seller_id))

    wb = load_workbook(file, data_only=True)
    ws = wb.active

    # Detect column layout by header row
    header_row = [str(c.value or "").strip().lower() for c in ws[1]]
    def _col(name):
        try:
            return header_row.index(name)
        except ValueError:
            return None

    ci_id     = _col("item_id")
    ci_marca  = _col("marca")
    ci_vid    = _col("variacao")
    ci_cost   = _col("custo_nf")
    ci_tax    = _col("aliquota_%")
    ci_fee    = _col("comissao_ml_%")
    ci_ship   = _col("frete_unit")

    if ci_id is None or ci_cost is None:
        flash("Planilha inválida — colunas 'item_id' e 'custo_nf' são obrigatórias.", "error")
        return redirect(url_for("costs.custos", seller_id=seller_id))

    # ── Phase 1: validate all rows ────────────────────────────────────────────
    validated = []
    row_errors = []

    for row_i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
        item_id = str(row[ci_id]).strip() if row[ci_id] else ""
        if not item_id or item_id.lower() == "none":
            continue

        errors = []

        def _num(idx, label, required=False):
            if idx is None:
                return 0.0
            val = row[idx]
            if val is None or val == "":
                if required:
                    errors.append(f"{label}: valor obrigatório")
                return 0.0
            try:
                f = float(val)
                if f < 0:
                    errors.append(f"{label}: não pode ser negativo (recebido {f})")
                return f
            except (ValueError, TypeError):
                errors.append(f"{label}: '{val}' não é um número")
                return 0.0

        unit_cost     = _num(ci_cost, "custo_nf")
        tax_pct       = _num(ci_tax,  "aliquota_%")
        ml_fee_pct    = _num(ci_fee,  "comissao_ml_%") or 14.0
        shipping      = _num(ci_ship, "frete_unit")
        marca         = str(row[ci_marca]).strip() if ci_marca is not None and row[ci_marca] else ""
        variation_id  = str(row[ci_vid]).strip()  if ci_vid  is not None and row[ci_vid]  else None

        if tax_pct > 100:
            errors.append(f"aliquota_%: {tax_pct} > 100%")
        if ml_fee_pct > 100:
            errors.append(f"comissao_ml_%: {ml_fee_pct} > 100%")

        if errors:
            row_errors.append({"linha": row_i, "item_id": item_id, "erros": errors})
        else:
            validated.append({
                "item_id": item_id, "variation_id": variation_id,
                "unit_cost": unit_cost, "tax_pct": tax_pct,
                "ml_fee_pct": ml_fee_pct, "shipping": shipping, "marca": marca,
            })

    # ── Phase 2: abort if any errors ─────────────────────────────────────────
    if row_errors:
        error_lines = "; ".join(
            f"L{e['linha']} ({e['item_id']}): {', '.join(e['erros'])}"
            for e in row_errors[:10]
        )
        if len(row_errors) > 10:
            error_lines += f" … e mais {len(row_errors) - 10} erros."
        flash(
            f"Importação cancelada — {len(row_errors)} erro(s) encontrado(s). "
            f"Nenhum dado salvo. Corrija e reimporte. // {error_lines}",
            "error",
        )
        return redirect(url_for("costs.custos", seller_id=seller_id))

    # ── Phase 3: save all ─────────────────────────────────────────────────────
    for v in validated:
        save_cost(
            seller_id, v["item_id"], v["variation_id"],
            v["unit_cost"], v["tax_pct"] / 100.0,
            v["ml_fee_pct"] / 100.0, v["shipping"], v["marca"],
        )

    flash(f"Importação concluída — {len(validated)} produto(s) salvos.", "success")
    return redirect(url_for("costs.custos", seller_id=seller_id))


# ── Importar NF-e (XML) ────────────────────────────────────────────────────────

@costs_bp.route("/custos/<seller_id>/import-nfe", methods=["POST"])
def import_nfe(seller_id):
    account = _require_seller(seller_id)
    if not account:
        flash("Conta não encontrada.", "error")
        return redirect(url_for("costs.custos", seller_id=seller_id))

    file = request.files.get("nfe_file")
    if not file or not file.filename.lower().endswith(".xml"):
        flash("Envie um arquivo .xml válido.", "error")
        return redirect(url_for("costs.custos", seller_id=seller_id))

    try:
        parsed = parse_nfe_xml(file.read())
    except NFeParseError as e:
        flash(f"Erro ao ler o XML da NF-e: {e}", "error")
        return redirect(url_for("costs.custos", seller_id=seller_id))

    try:
        doc_id = save_nfe_document(
            seller_id, parsed["chave_acesso"], parsed["supplier_cnpj"],
            parsed["supplier_name"], parsed["emitted_at"],
        )
    except ValueError as e:
        flash(str(e), "error")
        return redirect(url_for("costs.custos", seller_id=seller_id))

    n_items = save_nfe_items(doc_id, parsed["items"])
    n_auto = auto_reconcile_nfe_items(seller_id, doc_id)
    n_pending = n_items - n_auto

    msg = f"Nota importada — {n_items} item(ns) da nota de '{parsed['supplier_name'] or parsed['supplier_cnpj']}'."
    if n_auto:
        msg += f" {n_auto} conciliado(s) automaticamente (custo já atualizado)."
    if n_pending:
        msg += f" {n_pending} aguardando conciliação manual."
    flash(msg, "success")

    if n_pending:
        return redirect(url_for("costs.conciliacao_nfe", seller_id=seller_id))
    return redirect(url_for("costs.custos", seller_id=seller_id))


# ── Conciliação manual NF-e × anúncio ──────────────────────────────────────────

@costs_bp.route("/custos/<seller_id>/conciliacao")
def conciliacao_nfe(seller_id):
    account = _require_seller(seller_id)
    if not account:
        flash("Conta não encontrada.", "error")
        return redirect(url_for("web.index"))

    pending = get_pending_nfe_items(seller_id)
    items   = get_items_for_costs(seller_id)

    return render_template(
        "conciliacao.html",
        account   = account,
        seller_id = seller_id,
        pending   = pending,
        items     = items,
    )


@costs_bp.route("/custos/<seller_id>/conciliacao/<int:nfe_item_id>", methods=["POST"])
def conciliacao_nfe_save(seller_id, nfe_item_id):
    account = _require_seller(seller_id)
    if not account:
        flash("Conta não encontrada.", "error")
        return redirect(url_for("web.index"))

    item_id = request.form.get("item_id", "").strip()
    if not item_id:
        flash("Selecione um anúncio pra vincular.", "error")
        return redirect(url_for("costs.conciliacao_nfe", seller_id=seller_id))

    try:
        reconcile_nfe_item(nfe_item_id, seller_id, item_id)
        flash("Item conciliado e custo atualizado.", "success")
    except ValueError as e:
        flash(str(e), "error")

    return redirect(url_for("costs.conciliacao_nfe", seller_id=seller_id))


# ── Apuração por período ──────────────────────────────────────────────────────

@costs_bp.route("/apuracao/<seller_id>")
def apuracao(seller_id):
    account = _require_seller(seller_id)
    if not account:
        flash("Conta não encontrada.", "error")
        return redirect(url_for("web.index"))

    days = int(request.args.get("days", 30))
    days = max(1, min(days, 365))

    from app.services.analytics import compute_margin
    margin = compute_margin(seller_id, days)

    return render_template(
        "apuracao.html",
        account   = account,
        seller_id = seller_id,
        days      = days,
        margin    = margin,
    )
